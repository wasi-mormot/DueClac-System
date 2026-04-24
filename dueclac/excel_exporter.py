from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .config import OUTPUT_COLUMNS
from .models import ReportData


class ExcelReportExporter:
    def export(self, report: ReportData, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        if output_path.exists() and output_path.is_dir():
            raise ValueError(
                f"Output path points to a folder, not an Excel file: {output_path}"
            )
        if not output_path.suffix:
            output_path = output_path.with_suffix(".xlsx")
        output_path.parent.mkdir(parents=True, exist_ok=True)

        dataframe = pd.DataFrame(
            [record.to_output_row() for record in report.records],
        )
        dataframe = dataframe.reindex(columns=OUTPUT_COLUMNS[1:])

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._safe_sheet_title(report.title)

        total_columns = len(OUTPUT_COLUMNS)
        header_row = 3
        data_start_row = header_row + 1

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
        worksheet["A1"] = report.title
        worksheet["A1"].font = Font(size=14, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
        worksheet["A2"] = report.date_range
        worksheet["A2"].font = Font(size=11, italic=True)
        worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        total_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for column_index, header in enumerate(OUTPUT_COLUMNS, start=1):
            cell = worksheet.cell(row=header_row, column=column_index, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        if dataframe.empty:
            worksheet.cell(row=data_start_row, column=1, value="No traders matched the due rules.")
            worksheet.merge_cells(
                start_row=data_start_row,
                start_column=1,
                end_row=data_start_row,
                end_column=4,
            )
            worksheet.cell(row=data_start_row, column=1).alignment = Alignment(horizontal="left")
            worksheet.cell(row=data_start_row, column=1).font = Font(italic=True)
            data_end_row = data_start_row
        else:
            for row_offset, row_values in enumerate(dataframe.itertuples(index=False), start=0):
                current_row = data_start_row + row_offset
                serial_cell = worksheet.cell(
                    row=current_row,
                    column=1,
                    value=f"=ROW()-{header_row}",
                )
                serial_cell.border = thin_border
                serial_cell.alignment = Alignment(horizontal="center", vertical="center")

                for column_index, value in enumerate(row_values, start=2):
                    cell = worksheet.cell(row=current_row, column=column_index, value=value)
                    cell.border = thin_border
                    if column_index <= 3:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0.00"
            data_end_row = data_start_row + len(dataframe) - 1

        total_row = data_end_row + 1
        worksheet.cell(row=total_row, column=2, value="Total")
        if dataframe.empty:
            worksheet.cell(row=total_row, column=6, value="=0")
            worksheet.cell(row=total_row, column=7, value="=0")
        else:
            worksheet.cell(
                row=total_row,
                column=6,
                value=f"=SUBTOTAL(109,F{data_start_row}:F{data_end_row})",
            )
            worksheet.cell(
                row=total_row,
                column=7,
                value=f"=SUBTOTAL(109,G{data_start_row}:G{data_end_row})",
            )

        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=total_row, column=column_index)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = thin_border
            if column_index >= 6:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        worksheet.freeze_panes = "A4"
        if not dataframe.empty:
            worksheet.auto_filter.ref = f"A{header_row}:G{data_end_row}"

        column_widths = {
            "A": 5,
            "B": 25,
            "C": 18,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 14,
        }
        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

        for row_number in range(1, total_row + 1):
            worksheet.row_dimensions[row_number].height = 20
        worksheet.row_dimensions[1].height = 24
        worksheet.row_dimensions[2].height = 22

        try:
            workbook.save(output_path)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot save Excel file to '{output_path}'. "
                "Choose a writable .xlsx file path and make sure the file is not open in Excel."
            ) from exc
        return output_path

    def _safe_sheet_title(self, value: str) -> str:
        invalid = set(r'[]:*?/\\')
        sanitized = "".join(character for character in value if character not in invalid).strip()
        return sanitized[:31] or "Due Report"
